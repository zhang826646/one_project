
from apps.tasks.celery import celery_app,BaseTask
from common.libs.comm import now
import requests
import ujson
import time
import pandas
from openpyxl import load_workbook
import os
from qiniu import Auth, put_file, etag
from common.helper.qiniu_helper import update_file
import os

file_path=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))+'/调研_shz.xlsx'

url = 'http://gatewayapi.kotaiqiu.com/api/gateway'

headers={
    'user-agent': 'Dart/3.1 (dart:io)',
    'content-type': 'application/json',
    'accept-encoding': 'gzip',
    'content-length': '406',
    'host': 'gatewayapi.kotaiqiu.com',
    'Connection': 'keep-alive'
}

name_dic = {
    '4867':	'谈小娱-石家庄南马路店',
	'5719':'谈小娱 - 石家庄藁城北国店',
	'4294':'谈小娱 - 石家庄御景半岛店',
	'4483':'谈小娱 - 石家庄东商广场店',
	'4061':'谈小娱 - 石家庄盛和店',
	'3649':'谈小娱 - 石家庄铁道大学店',
	'4486':'谈小娱 - 石家庄谈固店',
	'4465':'谈小娱 - 石家庄上东城店',
	'5446':'谈小娱 - 石家庄谈固东街店',
	'3802':'谈小娱 - 石家庄建北华新路店',
	'5358':'谈小娱 - 石家庄华林国际店',
	'4919':'谈小娱 - 石家庄万达广场店',
	'5054':'谈小娱 - 石家庄理工学院店',
	'3539':'谈小娱 - 石家庄空中花园店',
	'3934':'谈小娱 - 石家庄永泰中街店',
	'4693':'谈小娱 - 石家庄塔谈店',
	'5181':'谈小娱 - 石家庄华兴街店',
	'5046':'谈小娱 - 石家庄常胜店'
}


_item = {
	'4867':'9c71e6ed1f020235d92049261f1aec0f',
	'5719':'9e41beedf7849612aaddf3da8cad77a3',
	'4294':'756f5dd072834ba11de71cbcbbcaf215',
	'4483':'392159aca28b2c8972f6fb34ee65f421',
	'4061':'51a84001a685b28cefe3769846b057a6',
	'3649':'c0f8321a9128cf992bfb3351a48f1e31',
	'4486':'93af40d12267719dda7965111ba99ea2',
	'4465':'f43b21ff626f45bab08a50c504c32010',
	'5446':'05e2697f11c686343bc8e3996b5ddc4a',
	'3802':'d12941d9b07941b7acd9a8ce9addf4cd',
	'5358':'fbec650b57378b2a8b21bc04a957bc45',
	'4919':'8d0245091293044e208b14d8104a2f76',
	'5054':'075fcde8d4f0fe2918134b6b578584a4',
	'3539':'9c26ce35a652bf676835bd96300985f4',
	'3934':'fb6d97986750133023e42c792dae5aa3',
	'4693':'1477c1d78082d9c0b24a2fd9e7d4c093',
	'5181':'f2489082326a62b9fa87d73810f7a08e',
	'5046':'90900b26f5aa1ea7c661bc6342029ebf',
}


@celery_app.task(bind=True)
async def buil_billiards_list(self: BaseTask):
    data = {
        "channelCode": "ios_api_get",
        "clientVersionCode": "280",
        "bizContent": "{\"sortDirection\":\"ASC\",\"sortType\":\"DISTANCE\",\"page\":{\"limit\":15,\"start\":1},\"latitude\":39.785954,\"longitude\":116.324669,\"cityId\":79}",
        "method": "com.yuyuka.billiards.api.authorized.new.billiards.rcmd.list",
        "timestamp": "1710001637969",
        "token": "012a2f45a66ae843e4e10f7b18c2e00a",
        "sign": "23e270897f624cc7eec29794b14b21aa"
    }
    data['timestamp'] = str(now())

    resp = requests.post(url,headers=headers,json=data)
    rep = resp.json()
    print(rep)
    billiards_list = ujson.loads(rep['bizContent'])

    _list = []
    columns = ['billiardsId', 'billiardsName', 'position', 'price', 'popularity', 'tags', 'distanceText', 'positionLongitude', 'positionLatitude', 'phoneNumber']
    print(billiards_list['items'])
    for value in billiards_list['items']:
        baseInfo = value['baseInfo']
        _list.append([value['billiardsId'], baseInfo['billiardsName'], baseInfo['position'],value['price'],value['popularity'],str(value['tags']), value['distanceText'], baseInfo['positionLongitude'], baseInfo['positionLatitude'], baseInfo['phoneNumber']])

    billiards_datafram = pandas.DataFrame(_list, columns=columns)
    print(billiards_datafram)
    billiards_datafram.to_excel(file_path, sheet_name='sheet1', index=False)
    wb = load_workbook(file_path)
    for i, name in name_dic.items():
        if name not in wb.sheetnames:
            wb.create_sheet(name)
    wb.save(file_path)
    await buil_billiards_list_2()
    return _list

async def buil_billiards_list_2():
    data = {
    "channelCode": "ios_api_get",
    "clientVersionCode": "280",
    "bizContent": "{\"sortDirection\":\"ASC\",\"sortType\":\"DISTANCE\",\"page\":{\"limit\":15,\"start\":2},\"latitude\":39.785954,\"longitude\":116.324669,\"cityId\":79}",
    "method": "com.yuyuka.billiards.api.authorized.new.billiards.rcmd.list",
    "timestamp": "1710002421679",
    "token": "012a2f45a66ae843e4e10f7b18c2e00a",
    "sign": "9a1717c9158f56732599810c4ab5f013"
    }
    data['timestamp'] = str(int(time.time()))

    resp = requests.post(url, headers=headers, json=data)
    rep = resp.json()
    print(rep)
    billiards_list = ujson.loads(rep['bizContent'])

    _list = []
    columns = ['billiardsId', 'billiardsName', 'position', 'price', 'popularity', 'tags', 'distanceText',
               'positionLongitude', 'positionLatitude', 'phoneNumber']
    print(billiards_list['items'])
    for value in billiards_list['items']:
        baseInfo = value['baseInfo']
        _list.append(
            [value['billiardsId'], baseInfo['billiardsName'], baseInfo['position'], value['price'], value['popularity'],
             str(value['tags']), value['distanceText'], baseInfo['positionLongitude'], baseInfo['positionLatitude'],
             baseInfo['phoneNumber']])

    billiards_datafram = pandas.DataFrame(_list, columns=columns)
    print(billiards_datafram)
    ll = pandas.read_excel(file_path, sheet_name='sheet1')
    print(ll.itertuples())
    with pandas.ExcelWriter(file_path, mode='a', engine="openpyxl",
                            if_sheet_exists='overlay') as writer:  # doctest: +SKIP
        if ll.itertuples():
            billiards_datafram.to_excel(writer, sheet_name='sheet1', index=False, startrow=len(ll) + 1, header=None)
        else:
            billiards_datafram.to_excel(writer, sheet_name='sheet1', index=False)



async def get_billiards_list(id,sign):
    req = {
        'channelCode': "ios_api_get",
        'clientVersionCode': "280",
        'bizContent': '{\"id\":'+f'{id}'+'}',
        'method': "com.yuyuka.billiards.api.authorized.user.table.list.query",
        'timestamp': str(now()),
        'token': "012a2f45a66ae843e4e10f7b18c2e00a",
        'sign': sign
    }
    resp = requests.post(url,headers=headers,json=req)
    rep = resp.json()
    billiards_list = ujson.loads(rep['bizContent'])

    _list = [time.strftime("%m-%d %H:%M:%S", time.localtime(time.time()))]
    columns = ['time']
    print(billiards_list)
    for value in billiards_list['items']:
        poolTables = value['poolTables']

        for i in poolTables:
            remainingTime = i.get('remainingTime')  # 到期时间
            _list.extend([f"{i['goodsId']}_{i['goodsName']}_{i['tableName']}",i['tableStatus']])
            columns.extend(['goodsId','tableStatus'])
    print(_list)
    print(columns)
    _list = [_list]
    billiards_datafram = pandas.DataFrame(_list, columns=columns)
    print(billiards_datafram,'!!!!')
    ll = pandas.read_excel(file_path, sheet_name=name_dic[id])
    print('行数',len(ll))

    with pandas.ExcelWriter(file_path,mode='a',engine = "openpyxl", if_sheet_exists='overlay') as writer:  # doctest: +SKIP
        if len(ll):
            billiards_datafram.to_excel(writer, sheet_name=name_dic[id], index=False, startrow=len(ll)+1, header=None)
        else:
            billiards_datafram.to_excel(writer, sheet_name=name_dic[id], index=False)

    print(ll)
    print(len(ll))
    return _list

@celery_app.task(bind=True)
async def update_execl(self: BaseTask):
    r = await update_file(f'file/调研_shz.xlsx', local_pth=file_path)
    return r


async def time_up():
    for id,value in name_dic.items():
        data_frame = pandas.read_excel(file_path,sheet_name=name_dic[id])
        print(data_frame)
        for index, row in data_frame.iterrows():
            timestamp = int(time.mktime(time.strptime('2023-'+data_frame.loc[index, 'time'], "%Y-%m-%d %H:%M:%S")))
            # print(timestamp, time.localtime(timestamp))
            # print(timestamp + 3600*8,time.localtime(timestamp + 3600*8))
            if timestamp < 1678672800:
                data_frame.loc[index, 'time'] = time.strftime("%m-%d %H:%M:%S",  time.localtime(timestamp + 3600*8))
        with pandas.ExcelWriter(file_path, mode='a', engine="openpyxl", if_sheet_exists='replace') as writer:  # doctest: +SKIP
                data_frame.to_excel(writer, sheet_name=name_dic[id], index=False)
        print(data_frame)
        # str_time = "2022-01-01 12:00:00"
        # timestamp = time.mktime(time.strptime(str_time, "%Y-%m-%d %H:%M:%S"))
        # print(timestamp)
        # print(data_frame)


@celery_app.task(bind=True)
async def w_billiards_list(self: BaseTask):
    for id,sign in _item.items():
        await get_billiards_list(id, sign)
        time.sleep(10)
