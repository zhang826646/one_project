
from apps.tasks.celery import celery_app,BaseTask
from common.libs.comm import now
import requests
import ujson
import time
import pandas
from openpyxl import load_workbook
import os
from qiniu import Auth, put_file, etag
from common.helper.qiniu_helper import update_file
import os

file_path=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))+'/调研.xlsx'

url = 'http://gatewayapi.kotaiqiu.com/api/gateway'

headers={
    'user-agent': 'Dart/3.1 (dart:io)',
    'content-type': 'application/json',
    'accept-encoding': 'gzip',
    'content-length': '406',
    'host': 'gatewayapi.kotaiqiu.com',
    'Connection': 'keep-alive'
}


_item = {
	'4549':'3859fb56d16937a9d587960ee903755b',
	'4118':'971fb59b83642261044f5eca9a0342f5',
	'5773':'517532c0576f2b510cb795b293d2f500',
	'3406':'be1aebec0f3f68f555e8c32df44314e7',
	'4591':'a1dda086e360fb5f2886448f54810fd8',
	'3628':'3d43fb3654a6a54883e44b0ef9bd7b4b',
	'3909':'1dab131f0d283ea14e3b75393fec8b5c',
	'3940':'45b31831117b0fc86cbb33c2b42bee17',
	'3362':'eded652016fddc5eea8f408a56440b48',
	'5066':'9f97e9e9512695e5765b3eadad1105e5',
	'3748':'12d37411d4c0f1c56dafa26c842d4e71',
	'3497':'fa7030109965bbc9a047da1f7589e88b',
	'5707':'12643bb768b0de31d316f8fc6e09a025',
	'3883':'c81642421b58b5b7aae43441e082ad82',
	'3800':'3cd2b7c990d5fbda866cb83f8d8ff2b0',
	'4703':'fe1e221950c7f7588079c95d60dadcad'
}

name_dic = {
	'4703': '谈小娱-邢台宁晋石坊路',
	'3800': '谈小娱-邢台南宫首店',
	'3883': '谈小娱-邢台唐尧街店',
	'5707': '谈小娱-邢台巨鹿建设北街店',
	'3497': '谈小娱-邢台太行北路店',
	'3748': '谈小娱-邢台上海城东区店',
	'5066': '谈小娱-邢台风荷曲苑店',
	'3362': '谈小娱-邢台泰和苑店',
	'3940': '谈小娱-邢台大都汇店',
	'3909': '谈小娱-威县中华大街店',
	'3628': '谈小娱-邢台任泽区店',
	'4591': '谈小娱-邢台建设大街店',
	'3406': '谈小娱-邢台燕云台店',
	'5773': '谈小娱-邢台万城店',
	'4118': '谈小娱-邢台钢铁路店',
	'4549': '谈小娱-邢台冶金路店'
}


@celery_app.task(bind=True)
async def buil_billiards_list(self: BaseTask):
    data = {
    'channelCode': "ios_api_get",
    'clientVersionCode': "280",
    'bizContent': "{\"sortDirection\":\"ASC\",\"sortType\":\"DISTANCE\",\"page\":{\"limit\":15,\"start\":1},\"latitude\":39.785954,\"longitude\":116.324669,\"cityId\":63}",
    'method': "com.yuyuka.billiards.api.authorized.new.billiards.rcmd.list",
    'timestamp': "1709864409341",
    'token': "012a2f45a66ae843e4e10f7b18c2e00a",
    'sign': "44396a1ff5bf9111bb6362ec4fa76467"
    }
    data['timestamp'] = str(now())

    resp = requests.post(url,headers=headers,json=data)
    rep = resp.json()
    print(rep)
    billiards_list = ujson.loads(rep['bizContent'])

    _list = []
    columns = ['billiardsId', 'billiardsName', 'position', 'price', 'popularity', 'tags', 'distanceText', 'positionLongitude', 'positionLatitude', 'phoneNumber']
    print(billiards_list['items'])
    for value in billiards_list['items']:
        baseInfo = value['baseInfo']
        _list.append([value['billiardsId'], baseInfo['billiardsName'], baseInfo['position'],value['price'],value['popularity'],str(value['tags']), value['distanceText'], baseInfo['positionLongitude'], baseInfo['positionLatitude'], baseInfo['phoneNumber']])

    billiards_datafram = pandas.DataFrame(_list, columns=columns)
    print(billiards_datafram)
    billiards_datafram.to_excel(file_path, sheet_name='sheet1', index=False)
    wb = load_workbook(file_path)
    for i, name in name_dic.items():
        if name not in wb.sheetnames:
            wb.create_sheet(name)
    wb.save(file_path)
    await _get_billiards_list_2()
    return _list


async def _get_billiards_list_2():
	data = {
	"channelCode": "ios_api_get",
	"clientVersionCode": "280",
	"bizContent": "{\"sortDirection\":\"ASC\",\"sortType\":\"DISTANCE\",\"page\":{\"limit\":15,\"start\":2},\"latitude\":39.785954,\"longitude\":116.324669,\"cityId\":63}",
	"method": "com.yuyuka.billiards.api.authorized.new.billiards.rcmd.list",
	"timestamp": "1710004853317",
	"token": "012a2f45a66ae843e4e10f7b18c2e00a",
	"sign": "38935080122a241070274cacdd158b59"
}
	data['timestamp'] = str(int(time.time()))

	resp = requests.post(url, headers=headers, json=data)
	rep = resp.json()
	print(rep)
	billiards_list = ujson.loads(rep['bizContent'])

	_list = []
	columns = ['billiardsId', 'billiardsName', 'position', 'price', 'popularity', 'tags', 'distanceText',
			   'positionLongitude', 'positionLatitude', 'phoneNumber']
	print(billiards_list['items'])
	for value in billiards_list['items']:
		baseInfo = value['baseInfo']
		_list.append(
			[value['billiardsId'], baseInfo['billiardsName'], baseInfo['position'], value['price'], value['popularity'],
			 str(value['tags']), value['distanceText'], baseInfo['positionLongitude'], baseInfo['positionLatitude'],
			 baseInfo['phoneNumber']])

	billiards_datafram = pandas.DataFrame(_list, columns=columns)
	print(billiards_datafram)
	ll = pandas.read_excel(file_path, sheet_name='sheet1')
	print(ll.itertuples())
	with pandas.ExcelWriter(file_path, mode='a', engine="openpyxl",
							if_sheet_exists='overlay') as writer:  # doctest: +SKIP
		if ll.itertuples():
			billiards_datafram.to_excel(writer, sheet_name='sheet1', index=False, startrow=len(ll) + 1, header=None)
		else:
			billiards_datafram.to_excel(writer, sheet_name='sheet1', index=False)


async def get_billiards_list(id,sign):
    req = {
        'channelCode': "ios_api_get",
        'clientVersionCode': "280",
        'bizContent': '{\"id\":'+f'{id}'+'}',
        'method': "com.yuyuka.billiards.api.authorized.user.table.list.query",
        'timestamp': str(now()),
        'token': "012a2f45a66ae843e4e10f7b18c2e00a",
        'sign': sign
    }
    resp = requests.post(url,headers=headers,json=req)
    rep = resp.json()
    billiards_list = ujson.loads(rep['bizContent'])

    _list = [time.strftime("%m-%d %H:%M:%S", time.localtime(time.time()))]
    columns = ['time']
    print(billiards_list)
    for value in billiards_list['items']:
        poolTables = value['poolTables']

        for i in poolTables:
            remainingTime = i.get('remainingTime')  # 到期时间
            _list.extend([f"{i['goodsId']}_{i['goodsName']}_{i['tableName']}",i['tableStatus']])
            columns.extend(['goodsId','tableStatus'])
    print(_list)
    print(columns)
    _list = [_list]
    billiards_datafram = pandas.DataFrame(_list, columns=columns)
    ll = pandas.read_excel(file_path, sheet_name=name_dic[id])
    print(ll.itertuples())
    with pandas.ExcelWriter(file_path,mode='a',engine = "openpyxl", if_sheet_exists='overlay') as writer:  # doctest: +SKIP
        if len(ll):
            billiards_datafram.to_excel(writer, sheet_name=name_dic[id], index=False, startrow=len(ll)+1, header=None)
        else:
            billiards_datafram.to_excel(writer, sheet_name=name_dic[id], index=False)

    print(ll)
    print(len(ll))
    return _list

@celery_app.task(bind=True)
async def update_execl(self: BaseTask):
    r = await update_file(f'file/调研.xlsx', local_pth=file_path)
    return r


@celery_app.task(bind=True)
async def update_execl(self: BaseTask):
    r = await update_file(f'file/调研.xlsx', local_pth=file_path)
    return r


async def time_up():
    for id,value in name_dic.items():
        data_frame = pandas.read_excel(file_path,sheet_name=name_dic[id])
        print(data_frame)
        for index, row in data_frame.iterrows():
            timestamp = int(time.mktime(time.strptime('2023-'+data_frame.loc[index, 'time'], "%Y-%m-%d %H:%M:%S")))
            # print(timestamp, time.localtime(timestamp))
            # print(timestamp + 3600*8,time.localtime(timestamp + 3600*8))
            if timestamp < 1678672800:
                data_frame.loc[index, 'time'] = time.strftime("%m-%d %H:%M:%S",  time.localtime(timestamp + 3600*8))
        with pandas.ExcelWriter(file_path, mode='a', engine="openpyxl", if_sheet_exists='replace') as writer:  # doctest: +SKIP
                data_frame.to_excel(writer, sheet_name=name_dic[id], index=False)
        print(data_frame)
        # str_time = "2022-01-01 12:00:00"
        # timestamp = time.mktime(time.strptime(str_time, "%Y-%m-%d %H:%M:%S"))
        # print(timestamp)
        # print(data_frame)

@celery_app.task(bind=True)
async def w_billiards_list(self: BaseTask):
    for id,sign in _item.items():
        await get_billiards_list(id, sign)
        time.sleep(10)
